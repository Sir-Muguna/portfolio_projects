import requests
import pandas as pd
import re
import os
import io
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from urllib.parse import quote  # Ensure URLs are safe

# Constants
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1TmO0yNbLGhvsQGu0Knxh7dd1h1SZOY2hIyOgxGgbLHI/export?format=csv"
DOWNLOAD_ROOT = '/root/projects/portfolio_projects/assent_ke_projects/nexperia'
HEADERS = {
    "User-Agent": "Mozilla/5.0"
}
BATCH_SIZE = 2000
SLEEP_SECONDS = 60
MAX_WORKERS = 10

# Ensure root folder exists
os.makedirs(DOWNLOAD_ROOT, exist_ok=True)

def process_manufacturer_pn(pn):
    if not isinstance(pn, str):
        return None
    if "," in pn:
        pn = pn.split(",")[0]
    elif "/" in pn[:-3] and re.search(r"[A-Z]$", pn):
        pn.strip()
    elif re.search(r"[A-Z]$", pn):
        pn = pn[:-1]
    return pn.strip()

def check_excel_file(file_path, original_pn):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if any(original_pn in str(cell) for cell in row if cell is not None):
                print(f"✅ Verified content for {original_pn}")
                return True
        print(f"❌ Content not found for {original_pn}")
    except Exception as e:
        print(f"❌ Error verifying {original_pn}: {e}")
    return False

def download_excel(pn, original_pn, url, folder_path):
    try:
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()
        file_path = os.path.join(folder_path, f"{pn}_Nexperia_Product_MDS.xlsx")
        with open(file_path, 'wb') as file:
            file.write(response.content)

        if check_excel_file(file_path, original_pn):
            print(f"📥 Downloaded: {pn}")
            return pn, original_pn, file_path, True
        else:
            os.remove(file_path)
            return pn, original_pn, None, False
    except requests.exceptions.RequestException as e:
        print(f"❌ Error downloading {pn}: {e}")
        return pn, original_pn, None, False

def main():
    try:
        response = requests.get(GOOGLE_SHEET_URL)
        response.raise_for_status()
        df = pd.read_csv(io.StringIO(response.content.decode('utf-8')))
        df['original_pn'] = df['Manufacturer PN']
        df['processed_pn'] = df['Manufacturer PN'].apply(process_manufacturer_pn)
        df = df.dropna(subset=['processed_pn'])
        df = df[['original_pn', 'processed_pn']]

        total_parts = len(df)
        num_batches = (total_parts + BATCH_SIZE - 1) // BATCH_SIZE
        print(f"🔢 Total parts: {total_parts}, Batches: {num_batches}")

        downloaded_files = []
        summary_records = []

        for batch_num in range(num_batches):
            start_idx = batch_num * BATCH_SIZE
            end_idx = min(start_idx + BATCH_SIZE, total_parts)
            batch_df = df.iloc[start_idx:end_idx].copy()

            batch_folder = os.path.join(DOWNLOAD_ROOT, f"batch_{batch_num + 1}")
            os.makedirs(batch_folder, exist_ok=True)

            print(f"\n🚚 Batch {batch_num + 1}: rows {start_idx} to {end_idx - 1}")

            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                futures = {
                    executor.submit(
                        download_excel,
                        row['processed_pn'],
                        row['original_pn'],
                        f"https://www.nexperia.com/.rest/chemical-content/v1/downloadExcel/{quote(row['processed_pn'])}",
                        batch_folder
                    ): row
                    for _, row in batch_df.iterrows()
                }

                for future in as_completed(futures):
                    row = futures[future]
                    try:
                        pn, original_pn, file_path, found = future.result()
                        summary_records.append({
                            'original_pn': original_pn,
                            'processed_pn': pn,
                            'file_path': file_path if file_path else "",
                            'found': found
                        })
                        if found and file_path and os.path.exists(file_path):
                            downloaded_files.append(file_path)
                    except Exception as e:
                        print(f"🚨 Error processing {row['processed_pn']}: {e}")
                        summary_records.append({
                            'original_pn': row['original_pn'],
                            'processed_pn': row['processed_pn'],
                            'found': False
                        })

            if batch_num < num_batches - 1:
                print("🕒 Waiting 2 minutes before next batch...")
                time.sleep(SLEEP_SECONDS)

        summary_df = pd.DataFrame(summary_records)
        summary_path = os.path.join(DOWNLOAD_ROOT, "successful_downloads_summary.xlsx")
        summary_df.to_excel(summary_path, index=False)
        downloaded_files.append(summary_path)

        zip_path = os.path.join(DOWNLOAD_ROOT, "downloaded_files.zip")
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file in downloaded_files:
                if file and os.path.exists(file):
                    zipf.write(file, os.path.basename(file))
                else:
                    print(f"⚠️ Missing file skipped: {file}")

        print(f"\n📦 All downloads completed. Archive saved to: {zip_path}")

    except Exception as e:
        print(f"🚨 Script error: {e}")

if __name__ == "__main__":
    main()
