import requests
import pandas as pd
import re
import os
import io
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook

# Constants
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1WR6PAJb3NVX0ynq-_SRJKhFWXQcrkXDKwGdioG-1Ilw/export?format=csv"
DOWNLOAD_ROOT = '/root/projects/portfolio_projects/assent_ke_projects/nexperia'
HEADERS = {"User-Agent": "Mozilla/5.0"}
BATCH_SIZE = 1000
SLEEP_SECONDS = 60
MAX_WORKERS = 10

# Ensure root folder exists
os.makedirs(DOWNLOAD_ROOT, exist_ok=True)

# Process Manufacturer PN
def process_manufacturer_pn(pn):
    if not isinstance(pn, str):
        return None
    if "," in pn:
        pn = pn.split(",")[0]
    elif re.search(r"[A-Z]$", pn):
        pn = pn[:-1]
    return pn.strip()

# Check content of the Excel file (Row 4 only)
def check_excel_file(file_path, original_pn):
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        print(f"🔍 Checking content for {original_pn} in file: {file_path}")
        row_4 = list(sheet.iter_rows(values_only=True))[3]  # Row 4 is index 3
        print(f"Row 4: {row_4}")
        if any(original_pn in str(cell) for cell in row_4 if cell is not None):
            print(f"✅ Found {original_pn} in file: {file_path}")
            return True
        else:
            print(f"❌ No matching content for {original_pn}. Skipped.")
    except Exception as e:
        print(f"❌ Error checking Excel file: {file_path}, Error: {e}")
    return False

# Download Excel file
def download_excel(pn, original_pn, url, folder_path):
    try:
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()

        file_path = os.path.join(folder_path, f"{pn}_Nexperia_Product_MDS.xlsx")
        with open(file_path, 'wb') as file:
            file.write(response.content)

        if check_excel_file(file_path, original_pn):
            print(f"✅ Downloaded and verified: {file_path}")
            return pn, original_pn, file_path
        else:
            os.remove(file_path)
            return pn, original_pn, None
    except requests.exceptions.RequestException as e:
        print(f"❌ Failed to download for PN: {pn}. Error: {e}")
        return pn, original_pn, None

# Main script
def main():
    try:
        # Step 1: Fetch the Google Sheet
        response = requests.get(GOOGLE_SHEET_URL)
        response.raise_for_status()
        df = pd.read_csv(io.StringIO(response.content.decode('utf-8')))

        # Step 2: Preprocess
        df['original_pn'] = df['Manufacturer PN']
        df['processed_pn'] = df['Manufacturer PN'].apply(process_manufacturer_pn)
        df = df.dropna(subset=['processed_pn'])
        df = df[['original_pn', 'processed_pn']]

        total_parts = len(df)
        num_batches = (total_parts + BATCH_SIZE - 1) // BATCH_SIZE
        print(f"🧾 Total parts: {total_parts}, Number of batches: {num_batches}")

        successful_parts = []
        downloaded_files = []

        for batch_num in range(num_batches):
            start_idx = batch_num * BATCH_SIZE
            end_idx = min(start_idx + BATCH_SIZE, total_parts)
            batch_df = df.iloc[start_idx:end_idx].copy()

            batch_folder = os.path.join(DOWNLOAD_ROOT, f"batch_{batch_num + 1}")
            os.makedirs(batch_folder, exist_ok=True)

            print(f"\n📦 Starting Batch {batch_num + 1}: rows {start_idx} to {end_idx - 1}")

            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                future_to_pn = {
                    executor.submit(
                        download_excel,
                        row['processed_pn'],
                        row['original_pn'],
                        f"https://www.nexperia.com/.rest/chemical-content/v1/downloadExcel/{row['processed_pn']}",
                        batch_folder
                    ): row['processed_pn']
                    for _, row in batch_df.iterrows()
                }

                for future in as_completed(future_to_pn):
                    result = future.result()
                    if result:
                        pn, original_pn, file_path = result
                        if file_path and os.path.exists(file_path):
                            successful_parts.append({'processed_pn': pn, 'original_pn': original_pn})
                            downloaded_files.append(file_path)

            if batch_num < num_batches - 1:
                print("⏳ Sleeping for 1 minute before next batch...")
                time.sleep(SLEEP_SECONDS)

        # Step 3: Save the summary of successfully downloaded part numbers
        summary_path = os.path.join(DOWNLOAD_ROOT, "successful_downloads.xlsx")
        pd.DataFrame(successful_parts).to_excel(summary_path, index=False)
        downloaded_files.append(summary_path)

        # Step 4: Create a zip archive with all downloaded files and the summary
        zip_path = os.path.join(DOWNLOAD_ROOT, "downloaded_files.zip")
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file in downloaded_files:
                if file and os.path.exists(file):
                    zipf.write(file, os.path.basename(file))
                else:
                    print(f"⚠️ Skipping missing file: {file}")

        print(f"\n✅ All downloaded files and summary saved to: {zip_path}")

    except Exception as e:
        print(f"🚨 An error occurred: {e}")

if __name__ == "__main__":
    main()
